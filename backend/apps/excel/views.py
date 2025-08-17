from backend.apps.space.models import Space, SpaceBackup
from backend.apps.account.permissions import IsSpaceMember
from backend.apps.history.models import HistoryIncome, HistoryExpense
from backend.apps.total_balance.models import TotalBalance
from backend.apps.total_balance.serializers import TotalBalanceSerializer
from backend.apps.category.serializers import CategorySerializer
from backend.apps.account.models import Account
from backend.apps.category.models import Category
from backend.apps.converter.utils import convert_currencies
from backend.apps.account.serializers import AccountSerializer
from backend.apps.converter.utils import convert_number_to_letter

from rest_framework.response import Response
from datetime import datetime
from django.utils import timezone
from django.utils.timezone import make_aware, localtime
from dateutil.parser import parse
from django.http import HttpResponse
from openpyxl.styles import Font
from rest_framework import generics, status
from django.db import transaction
from decimal import Decimal, ROUND_DOWN

import pytz
import calendar
import json
import openpyxl
import random


class DecimalEncoder(json.JSONEncoder):
    def default(self, obj):
        if isinstance(obj, Decimal):
            return float(obj)
        return super().default(obj)

def get_random_color():
    presets = [
        "#FF9800", 
        "#4CAF50", 
        "#2196F3", 
        "#9C27B0", 
        "#F44336",  
        "#00BCD4",
        "#FF7DD1",
        "#795548", 
        "#607D8B", 
        "#FFEC42", 
        "#3448D4",
        "#47AE0D",
    ]
    return random.choice(presets)

class ExportHistoryView(generics.GenericAPIView):
    permission_classes = (IsSpaceMember,)

    @staticmethod
    def post(request, *args, **kwargs):
        highest_role = request.user.roles[0]
        if highest_role == 'free':
            return Response("Error: You are not allowed to export history data with a free role.", status=status.HTTP_403_FORBIDDEN)

        space = Space.objects.get(pk=kwargs['space_pk'])
        print(request.data)
        expenses = HistoryExpense.objects.filter(father_space=space)
        incomes = HistoryIncome.objects.filter(father_space=space)

        filters = request.data.get("filters", {})
        if "dates" in filters:
            tz = pytz.timezone('UTC')
            start_date = request.data.get("from_date")
            end_date = request.data.get("to_date")
            if start_date and end_date:
                start_date = make_aware(datetime.strptime(start_date, '%Y-%m-%d'), tz)
                end_date = make_aware(datetime.strptime(end_date, '%Y-%m-%d'), tz)
                expenses = expenses.filter(created__range=(start_date, end_date))
                incomes = incomes.filter(created__range=(start_date, end_date))

        # Фильтрация по счетам
        if filters.get("accounts") is True:
            account_ids = filters.get("account_ids", [])
            if account_ids:
                expenses = expenses.filter(from_acc__id__in=account_ids, transaction_id__isnull=True)
                incomes = incomes.filter(account__id__in=account_ids, transaction_id__isnull=True)

        # Фильтрация по картам
        if filters.get("cards", False):
            card_ids = filters.get("cards_ids", [])
            if card_ids:
                # Фильтруем только карты (transaction_id__isnull=False)
                card_expenses = HistoryExpense.objects.filter(
                    father_space=space,
                    from_acc__id__in=card_ids,
                    transaction_id__isnull=False  # Только карты
                )
                card_incomes = HistoryIncome.objects.filter(
                    father_space=space,
                    account__id__in=card_ids,
                    transaction_id__isnull=False  # Только карты
                )
                # Если accounts тоже True, объединяем результаты
                if filters.get("accounts", False):
                    expenses = expenses | card_expenses
                    incomes = incomes | card_incomes
                else:
                    # Если только cards, то используем только записи по картам
                    expenses = card_expenses
                    incomes = card_incomes

        export_type = request.data.get("export_type", "both")

        if export_type == "expenses":
            incomes = []
        elif export_type == "incomes":
            expenses = []

        # Определяем заголовки
        default_headers = {
            'amount': 'Amount',
            'comment': 'Comment',
            'date': 'Date & Time',
            'account': 'Account',
            'category': 'Category'
        }

        # Если export_type == 'both', добавляем колонку 'type' в начало
        if export_type == "both":
            default_headers = {'type': 'Type', **default_headers}

        fields = request.data.get("fields", list(default_headers.keys()))
        headers = [default_headers[field] for field in fields if field in default_headers]

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "History Data"

        ws.append(headers)

        for cell in ws[1]:
            cell.font = Font(bold=True)

        for expense in expenses:
            row = []
            for field in fields:
                try:
                    cat_title, cat_icon, history_type = expense.to_cat["title"], expense.to_cat["icon"], "expense"
                except TypeError:
                    cat_title, cat_icon, history_type = "", "", "loss"
                if field == 'type' and export_type == "both":
                    row.append(history_type)
                elif field == 'amount':
                    row.append(expense.amount)
                elif field == 'comment':
                    row.append(expense.comment or '')
                elif field == 'date':
                    row.append(localtime(expense.created).strftime('%d %b %Y @ %H:%M'))
                elif field == 'account':
                    row.append(expense.from_acc.get('title', '') if expense.from_acc else '')
                elif field == 'category':
                    row.append(expense.to_cat.get('title', '') if expense.to_cat else '')
            ws.append(row)

        for income in incomes:
            row = []
            for field in fields:
                if field == 'type' and export_type == "both":
                    row.append('income')
                elif field == 'amount':
                    row.append(income.amount)
                elif field == 'comment':
                    row.append(income.comment or '')
                elif field == 'date':
                    row.append(localtime(income.created).strftime('%d %b %Y @ %H:%M'))
                elif field == 'account':
                    row.append(income.account.get('title', '') if income.account else '')
                elif field == 'category':
                    row.append('')
            ws.append(row)

        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response['Content-Disposition'] = 'attachment; filename="history_data.xlsx"'
        wb.save(response)
        return response


class ImportHistoryView(generics.GenericAPIView):
    permission_classes = (IsSpaceMember,)

    @staticmethod
    def post(request, *args, **kwargs):
        highest_role = request.user.roles[0]
        if highest_role == 'free':
            return Response("Error: You are not allowed to import history data with a free role.", status=status.HTTP_403_FORBIDDEN)
        
        space_id = kwargs.get("space_pk")
        try:
            father_space = Space.objects.get(pk=space_id)
        except Space.DoesNotExist:
            return Response(
                {"error": "Space with the given ID does not exist."},
                status=status.HTTP_404_NOT_FOUND
            )

        uploaded_file = request.FILES.get("file")
        record_type = request.data.get("type")
        account_id = request.data.get("account_id")

        if not uploaded_file or not record_type or not account_id:
            return Response(
                {"error": "file, type, and account_id are required."},
                status=status.HTTP_400_BAD_REQUEST
            )

        if record_type not in ["expense", "income"]:
            return Response(
                {"error": "Invalid type. Must be 'expense' or 'income'."},
                status=status.HTTP_400_BAD_REQUEST
            )

        account = Account.objects.filter(pk=account_id, father_space=father_space).first()
        if not account:
            return Response({"error": "Account not found"}, status=404)

        try:
            # Валидация файла: проверка, что это действительно Excel-файл
            wb = openpyxl.load_workbook(uploaded_file, read_only=False, data_only=True)

            if len(wb.sheetnames) > 1:
                return Response(
                    {"error": "The file must contain only one worksheet."},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            ws = wb.active
        except Exception as e:
            return Response(
                {"error": "Invalid file format. Please upload a valid Excel file."},
                status=status.HTTP_400_BAD_REQUEST
            )

        if ws.max_row is None or ws.max_row < 2:
            print(ws.max_row)
            return Response(
                {"error": "The file must contain at least one row of data."},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Определяем индексы колонок для обязательных полей
        try:
            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        except StopIteration:
            return Response(
                {"error": "The file is empty or does not contain any data."},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        column_indices = {
            'amount': None,
            'created': None,
            'category': None,
            'comment': None
        }

        for idx, header in enumerate(header_row):
            if header and isinstance(header, str):
                header_lower = header.strip().lower()
                if header_lower == 'amount':
                    column_indices['amount'] = idx
                elif header_lower in ['date', 'created']:
                    column_indices['created'] = idx
                elif header_lower == 'category': 
                    column_indices['category'] = idx
                elif header_lower == 'comment':
                    column_indices['comment'] = idx

        if column_indices['amount'] is None or column_indices['created'] is None or column_indices['category'] is None:
            return Response(
                {"error": "The file must contain 'amount', 'created' and 'category' columns."},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Предварительная валидация всех строк
        validation_errors = []
        rows_data = []
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            # Проверяем, что строка не пустая
            if not any(cell for cell in row if cell is not None):
                continue  # Пропускаем пустые строки
                
            row_data = {}
            errors = []
            
            try:
                # Валидация amount
                amount_raw = row[column_indices['amount']] if column_indices['amount'] < len(row) else None
                if amount_raw is None or amount_raw == '':
                    errors.append(f"Missing amount in row {row_idx}")
                else:
                    try:
                        amount_flo = abs(float(amount_raw))
                        amount = Decimal(str(amount_flo)).quantize(Decimal('0.01'), rounding=ROUND_DOWN)
                        if amount == 0:
                            errors.append(f"Amount cannot be zero in row {row_idx}")
                        row_data['amount'] = amount
                    except (ValueError, TypeError) as e:
                        errors.append(f"Invalid amount format in row {row_idx}: {amount_raw}")

                # Валидация category
                category_raw = row[column_indices['category']] if column_indices['category'] < len(row) else None
                if not category_raw or str(category_raw).strip() == '':
                    errors.append(f"Missing category in row {row_idx}")
                else:
                    row_data['category_name'] = str(category_raw).strip()[:24]

                # Валидация created
                created_raw = row[column_indices['created']] if column_indices['created'] < len(row) else None
                if created_raw is None:
                    errors.append(f"Missing date in row {row_idx}")
                else:
                    try:
                        if isinstance(created_raw, str):
                            created_dt = parse(created_raw)
                            created = make_aware(created_dt)
                        elif isinstance(created_raw, datetime):
                            created = make_aware(created_raw)
                        else:
                            errors.append(f"Invalid date format in row {row_idx}: {created_raw}")
                            continue
                        row_data['created'] = created
                    except (ValueError, TypeError) as e:
                        errors.append(f"Invalid date format in row {row_idx}: {created_raw}")

                # Получаем comment если есть
                if column_indices['comment'] is not None and column_indices['comment'] < len(row):
                    row_data['comment'] = str(row[column_indices['comment']]) if row[column_indices['comment']] else ''
                else:
                    row_data['comment'] = ''

                row_data['row_idx'] = row_idx
                
            except IndexError:
                errors.append(f"Row {row_idx} has insufficient columns")
            
            if errors:
                validation_errors.extend(errors)
            else:
                rows_data.append(row_data)

        # Если есть ошибки валидации, возвращаем их все
        if validation_errors:
            return Response(
                {"error": validation_errors},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Если валидация прошла успешно, выполняем импорт в транзакции
        try:
            with transaction.atomic():
                for row_data in rows_data:
                    amount = row_data['amount']
                    category_name = row_data['category_name']
                    created = row_data['created']
                    comment = row_data['comment']
                    
                    # Создаем запись в базе данных
                    if record_type == "expense":
                        if created.year == timezone.now().year and created.month == timezone.now().month:
                            category, created_category = Category.objects.get_or_create(
                                title=category_name,
                                father_space=father_space,
                                defaults={
                                    'spent': 0,
                                    'limit': None,
                                    'color': get_random_color(),
                                    'icon': 'Folder'
                                }
                            )

                            category.spent += amount
                            category.save()

                            # данные берем из объекта Category
                            to_cat_data = {
                                'id': category.id,
                                'icon': category.icon,
                                'color': category.color,
                                'limit': float(category.limit) if category.limit is not None else 0.0,
                                'spent': float(category.spent),
                                'title': category.title,
                                'father_space': category.father_space.id
                            }

                        else:
                            backup = SpaceBackup.objects.filter(
                                father_space=father_space,
                                date__year=created.year,
                                date__month=created.month
                            ).first()

                            if backup:
                                backup_categories = backup.categories
                                category = None

                                for category_data in backup_categories:
                                    if category_data.get('title', '').strip() == category_name:
                                        current_spent = float(category_data.get('spent', 0))
                                        category_data['spent'] = current_spent + float(amount)
                                        category = category_data
                                        break

                                if not category:
                                    category = {
                                        'title': category_name,
                                        'spent': float(amount),
                                        'limit': None,
                                        "limit_formatted": None,
                                        "spent_percentage": None,
                                        'color': get_random_color(),
                                        'father_space': father_space.id,
                                        'icon': 'Folder'
                                    }
                                    backup_categories.append(category)

                                total_balance = backup.total_balance
                                if total_balance:
                                    current_expenses_str = total_balance.get('total_expenses_this_month', 0)
                                    current_expenses = float(current_expenses_str)
                                    total_balance['total_expenses_this_month'] = current_expenses + float(amount)

                                backup.categories = backup_categories
                                backup.total_balance = total_balance
                                backup.save()
                            else:
                                # Бэкап не найден, создаем новый
                                last_day = calendar.monthrange(created.year, created.month)[1]
                                backup_date = created.date().replace(day=last_day)

                                accounts = []

                                category = {
                                    'title': category_name,
                                    'spent': float(amount),
                                    'limit': None,
                                    "limit_formatted": None,
                                    "spent_percentage": None,
                                    'color': get_random_color(),
                                    'father_space': father_space.id,
                                    'icon': 'Folder'
                                }

                                categories = [category]

                                total_balance_data = {
                                    'balance': 0.00,
                                    "currency": father_space.currency,
                                    'formatted_balance': convert_number_to_letter(float(amount)),
                                    'total_expenses_this_month': float(amount),
                                    'total_income_this_month': 0.00
                                }

                                SpaceBackup.objects.create(
                                    father_space=father_space,
                                    date=backup_date,
                                    accounts=accounts,
                                    categories=categories,
                                    total_balance=total_balance_data
                                )

                            # данные берем из словаря backup
                            to_cat_data = {
                                'id': category.get('id'),
                                'icon': category.get('icon'),
                                'color': category.get('color'),
                                'limit': float(category.get('limit')) if category.get('limit') is not None else 0.0,
                                'spent': float(category.get('spent', 0)),
                                'title': category.get('title'),
                                'father_space': category.get('father_space')
                            }

                        # создаем HistoryExpense
                        HistoryExpense.objects.create(
                            amount=amount,
                            currency=account.currency,
                            amount_in_default_currency=convert_currencies(
                                from_currency=account.currency,
                                amount=amount,
                                to_currency=father_space.currency),
                            comment=comment,
                            from_acc={"id": account.id, "title": account.title, "balance": str(account.balance)},
                            to_cat=to_cat_data,
                            father_space=father_space,
                            created=created
                        )

                    elif record_type == "income":
                        if created.year == timezone.now().year and created.month == timezone.now().month:
                            current_account, created_account = Account.objects.get_or_create(
                                title=category_name,
                                father_space=father_space,
                                defaults={
                                    'balance': 0,
                                    'currency': father_space.currency,
                                    'included_in_total_balance': True
                                }
                            )

                            current_account.balance += amount
                            current_account.save()

                            total_balance, created_total = TotalBalance.objects.get_or_create(
                                father_space=father_space,
                                defaults={'balance': 0}
                            )
                            total_balance.balance += amount
                            total_balance.save()

                            account = current_account
                            new_balance = current_account.balance

                        else:
                            backup = SpaceBackup.objects.filter(
                                father_space=father_space,
                                date__year=created.year,
                                date__month=created.month
                            ).first()
                            
                            if backup:
                                backup_accounts = backup.accounts
                                account = None
                                account_found = False

                                for account_data in backup_accounts:
                                    if account_data.get('title', '').strip() == category_name:
                                        current_balance = float(account_data.get('balance', 0))
                                        account_data['balance'] = current_balance + float(amount)
                                        account = account_data 
                                        account_found = True
                                        break
                                
                                if not account_found:
                                    new_account = {
                                        'title': category_name,
                                        "spend": 0,
                                        "formatted_spend": 0,
                                        "income": float(amount),
                                        "formatted_income": 0,
                                        'balance': float(amount),
                                        "balance_converted": convert_number_to_letter(float(amount)),
                                        'currency': father_space.currency,
                                        'father_space': father_space.id,
                                        'included_in_total_balance': True
                                    }
                                    backup_accounts.append(new_account)
                                    account = new_account  # назначаем новый аккаунт
                                
                                total_balance = backup.total_balance
                                if total_balance:
                                    current_income = total_balance.get('total_income_this_month', 0)
                                    balance = total_balance.get('balance', 0)
                                    if isinstance(current_income, str):
                                        current_income = float(current_income)
                                    if isinstance(balance, str):
                                        balance = float(balance)
                                    total_balance['total_income_this_month'] = current_income + float(amount)
                                    total_balance['balance'] = balance + float(amount)
                                
                                backup.accounts = backup_accounts
                                backup.total_balance = total_balance
                                backup.save()
                                new_balance = account['balance'] if isinstance(account, dict) else account.balance

                            else:
                                # Бэкап не найден, создаем новый
                                last_day = calendar.monthrange(created.year, created.month)[1]
                                backup_date = created.date().replace(day=last_day)
                                
                                new_account = {
                                    'title': category_name,
                                    "spend": 0,
                                    "formatted_spend": 0,
                                    "income": float(amount),
                                    "formatted_income": convert_number_to_letter(float(amount)),
                                    'balance': float(amount),
                                    "balance_converted": convert_number_to_letter(float(amount)),
                                    'currency': father_space.currency,
                                    'father_space': father_space.id,
                                    'included_in_total_balance': True
                                }
                                
                                accounts = [new_account]
                                
                                total_balance_data = {
                                    'balance': float(amount),
                                    "currency": father_space.currency,
                                    'formatted_balance': convert_number_to_letter(float(amount)),
                                    'total_expenses_this_month': 0.00,
                                    'total_income_this_month': float(amount)
                                }

                                account = new_account
                                new_balance = new_account['balance']
                                
                                SpaceBackup.objects.create(
                                    father_space=father_space,
                                    date=backup_date,
                                    accounts=accounts,
                                    categories=[],
                                    total_balance=total_balance_data
                                )

                        HistoryIncome.objects.create(
                            amount=amount,
                            new_balance=new_balance,
                            currency=account['currency'] if isinstance(account, dict) else account.currency,
                            amount_in_default_currency=convert_currencies(
                                from_currency=account['currency'] if isinstance(account, dict) else account.currency,
                                amount=amount,
                                to_currency=father_space.currency
                            ),
                            comment=comment,
                            account={"id": account.get('id', None) if isinstance(account, dict) else account.id,
                                    "title": account['title'] if isinstance(account, dict) else account.title,
                                    "balance": str(account['balance'] if isinstance(account, dict) else account.balance)},
                            father_space=father_space,
                            created=created
                        )


                return Response({"message": "Import completed successfully."}, status=status.HTTP_200_OK)
                
        except Exception as e:
            print(f"Error during import: {str(e)}")
            return Response(
                {"error": f"Error during import: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

class PreviewHistoryView(generics.GenericAPIView):
    permission_classes = (IsSpaceMember,)

    @staticmethod
    def post(request, *args, **kwargs):
        space = Space.objects.get(pk=kwargs['space_pk'])
        expenses = HistoryExpense.objects.filter(father_space=space)
        incomes = HistoryIncome.objects.filter(father_space=space)

        filters = request.data.get("filters", {})
        if "dates" in filters:
            tz = pytz.timezone('UTC')
            start_date = request.data.get("from_date")
            end_date = request.data.get("to_date")
            if start_date and end_date:
                start_date = make_aware(datetime.strptime(start_date, '%Y-%m-%d'), tz)
                end_date = make_aware(datetime.strptime(end_date, '%Y-%m-%d'), tz)
                expenses = expenses.filter(created__range=(start_date, end_date))
                incomes = incomes.filter(created__range=(start_date, end_date))

        if "accounts" in filters:
            account_ids = request.data.get("account_ids", [])
            if account_ids:
                expenses = expenses.filter(from_acc__id__in=account_ids)
                incomes = incomes.filter(account__id__in=account_ids)

        preview_data = []

        for expense in expenses[:5]:
            preview_data.append({
                "type": "expense",
                "amount": float(expense.amount),
                "comment": expense.comment or "",
                "date": localtime(expense.created).strftime('%Y-%m-%d %H:%M:%S'),
                "account": expense.from_acc.title if expense.from_acc else "",
                "category": expense.to_cat.title if expense.to_cat else ""
            })

        for income in incomes[:5]:
            preview_data.append({
                "type": "income",
                "amount": float(income.amount),
                "comment": income.comment or "",
                "date": localtime(income.created).strftime('%Y-%m-%d %H:%M:%S'),
                "account": income.account.title if income.account else "",
                "category": ""
            })

        # Ограничиваем общий результат до 10 записей
        preview_data = preview_data[:10]

        return Response(preview_data, status=200)
